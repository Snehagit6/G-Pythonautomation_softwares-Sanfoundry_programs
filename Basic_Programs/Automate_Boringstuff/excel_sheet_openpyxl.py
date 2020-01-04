import openpyxl
from openpyxl.styles import Color,Font
from openpyxl.styles.colors import GREEN
class Spec:

    def __init__(self,sheet_name):
        self.sheet_name = sheet_name
        wb = openpyxl.load_workbook(self.sheet_name)
        # sheet=wb.get_sheet_by_name("Data")
        self.sheet = wb.active

    def data(self,func):

        list_header = [self.sheet.cell(row=1, column=j).value for j in range(1, self.sheet.max_column+1)]
# for i in range(2,4):
#     for j in range(1,sheet.max_column+1):
#         list.append(sheet.cell(row=i,column=j).value)
# #         (dict(zip(list_header,list)))
#

#     print(dict(zip(list_header,list)))
#     list.clear()
# #Enhancing using list comprehension
        def inner(param,**kwargs):
            cont=[]
            contents=[]
            for i in range(2, Spec.sheet.max_row + 1):
                for j in range(1, Spec.sheet.max_column+1):
                    val=Spec.sheet.cell(row=i, column=j).value
                    cont.append(val)

                contents.append(cont)
                cont=[]
        # rows=[[sheet.cell(row=i,column=j).value for i in range(2,sheet.max_column+1)for j in range(1,sheet.max_column+1)]]

            records=[dict(zip(list_header,record)) for record in contents]
            course=[r for r in records if str(r['Course'].strip(' '))=='Functions and OOPS']

            param=course[0]
            func(param)

        return inner

#Entire set of data with rows
    def display_data(self):
        for i in range(1,Spec.sheet.max_row+1):
            for j in range(1,Spec.sheet.max_column+1):
                print(Spec.sheet.cell(row=i,column=j).value,end='\t\t\t')
            print('\n')


# for row_cells in sheet.iter_rows():
#     for cell in row_cells:
#         if cell.value=='Python':
#             cell.font=Font(bold=True,color='FFFF3030')
#             print("Color changed successfully")
#     for c in(1,sheet.max_column+1):
#         cell_value=sheet.cell(row=r,column=c).value
#         print(cell_value)
#         if cell_value=='Python':
#             sheet.cell(row=r,column=c).font=Font(color=DARKGREEN)

#
# wb.save("G://Pythonexceldata.xlsx")
# wb.close()